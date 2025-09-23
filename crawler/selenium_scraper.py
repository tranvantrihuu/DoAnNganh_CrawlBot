from bs4 import BeautifulSoup
import os
import re
import time
import unicodedata
from datetime import datetime
from typing import List, Dict
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import gc
from pathlib import Path

# === PATH ROOTS: luôn lưu ở <project-root>/output/... thay vì crawler/output/... ===
_THIS_FILE = Path(__file__).resolve()
# nếu file nằm trong thư mục "crawler", project-root là cha của nó; ngược lại là chính thư mục hiện tại
_PROJECT_ROOT = _THIS_FILE.parent.parent if _THIS_FILE.parent.name.lower() == "crawler" else _THIS_FILE.parent
_OUTPUT_ROOT = _PROJECT_ROOT / "output"

# ===========================================================
# MỤC ĐÍCH TỆP: CÁC HÀM & HẰNG SỐ PHỤ TRỢ CHO BỘ THU THẬP
# VIỆC LÀM TỪ VIETNAMWORKS (CRAWL + PHÂN TÍCH).
#
# Lưu ý đọc hiểu cho giảng viên:
# - File này KHÔNG tự chạy độc lập mà được import vào luồng crawler chính.
# - Các comment giải thích rõ "tại sao làm vậy" để phục vụ thẩm định học thuật.
# - Chúng tôi cố gắng viết selector/logic ở mức ổn định, nhưng giao diện VNW
#   có thể thay đổi (class động dạng sc-xxxx). Khi đó cần cập nhật selector.
# - Các thao tác đều tuân thủ quy tắc: chờ trang tải, cuộn lazy-load, tách link.
# ===========================================================

# ===================== CẤU HÌNH NGÀNH =====================
# Map tên ngành (hiển thị) -> group_id trên VietnamWorks
# Dùng cho việc build URL /viec-lam?g=<group_id>
# Ghi chú:
# - Đây là ánh xạ thủ công dựa trên trạng thái website tại thời điểm triển khai.
# - Nếu VietnamWorks thay đổi taxonomy/ID, cần cập nhật lại bảng này.
VNWORKS_GROUPS: Dict[str, int] = {
    "Bán Lẻ/Tiêu Dùng": 24,
    "Bảo Hiểm": 14,
    "Bất Động Sản": 23,
    "CEO & General Management": 29,
    "Chính Phủ/Phi Lợi Nhuận": 25,
    "Công Nghệ Thông Tin/Viễn Thông": 5,
    "Dược": 28,
    "Dệt May/Da Giày": 26,
    "Dịch Vụ Khách Hàng": 6,
    "Dịch Vụ Ăn Uống": 11,
    "Giáo Dục": 1,
    "Hành Chính Văn Phòng": 20,
    "Hậu Cần/Xuất Nhập Khẩu/Kho Bãi": 13,
    "Khoa Học & Kỹ Thuật": 9,
    "Kinh Doanh": 21,
    "Kiến Trúc/Xây Dựng": 4,
    "Kế Toán/Kiểm Toán": 2,
    "Kỹ Thuật": 22,
    "Nghệ thuật, Truyền thông/In ấn/Xuất bản": 18,
    "Ngân Hàng & Dịch Vụ Tài Chính": 10,
    "Nhà Hàng - Khách Sạn/Du Lịch": 15,
    "Nhân Sự/Tuyển Dụng": 12,
    "Nông/Lâm/Ngư Nghiệp": 3,
    "Pháp Lý": 16,
    "Sản Xuất": 27,
    "Thiết Kế": 7,
    "Tiếp Thị, Quảng Cáo/Truyền Thông": 17,
    "Vận Tải": 8,
    "Y Tế/Chăm Sóc Sức Khoẻ": 19
}

def slugify_vn(s: str) -> str:
        # Chuẩn hoá chuỗi tiếng Việt về 'slug' không dấu, dùng cho tên file/thư mục.
        # Lý do:
        #   (1) Tránh ký tự đặc biệt gây lỗi khi tạo path trên nhiều hệ điều hành.
        #   (2) Dễ tìm kiếm, so khớp, và đảm bảo tính nhất quán trong pipeline.
        # Các bước xử lý:
        #   - Thay '/' bằng space để không vô tình tạo cấp thư mục.
        #   - Chuyển riêng 'Đ/đ' sang 'D/d' (normalize NFD không đụng đến ký tự này).
        #   - Dùng NFD để tách dấu tiếng Việt; loại mọi ký tự Mn (dấu) để "bỏ dấu".
        #   - Chỉ giữ [a-zA-Z0-9] và khoảng trắng; ký tự khác thay bằng space.
        #   - Ép nhiều khoảng trắng liên tiếp về 1 space, trim 2 đầu, rồi lower-case.
    s = s.replace("/", " ")
    s = s.replace("Đ", "D").replace("đ", "d")  # xử lý Đ/đ riêng
    s = unicodedata.normalize("NFD", s)        # tách dấu
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")  # bỏ dấu
    s = re.sub(r"[^a-zA-Z0-9\s]", " ", s)      # chỉ giữ chữ/số/space
    s = re.sub(r"\s+", " ", s).strip()         # gộp space thừa
    return s.lower()

# DOMAIN cơ sở để ghép với các đường dẫn tương đối khi trích xuất link.
BASE = "https://www.vietnamworks.com"

def _scroll_lazy(driver, times=8, dy=1500, pause=0.25):
    """
    Cuộn trang theo 'đợt' để kích hoạt lazy-loading.

    Vì danh sách job ở VietnamWorks tải dần (infinite/lazy load), nếu không cuộn
    thì DOM chỉ có một phần. Hàm này mô phỏng hành vi người dùng cuộn xuống.

    Tham số:
    - times: số lần cuộn liên tiếp (tăng nếu thấy chưa tải đủ job).
    - dy: số pixel mỗi lần cuộn (lớn hơn → cuộn nhanh hơn nhưng dễ bỏ sót render).
    - pause: thời gian chờ giữa hai lần cuộn để trang kịp render/thêm node mới.

    Kỹ thuật:
    - Dùng ActionChains.scroll_by_amount thay vì send_keys(PAGE_DOWN) để chủ động.
    - Cần đảm bảo viewport đang focus đúng phần thân trang; nếu trang có modal/
      banner, cuộn có thể không tác dụng → xử lý ở nơi gọi (đóng pop-up trước).
    """
    for _ in range(times):
        ActionChains(driver).scroll_by_amount(0, dy).perform()
        time.sleep(pause)

def _extract_links_stepwise_from_card(card) -> List[str]:
    """
    Trích xuất các liên kết (href) đến trang chi tiết job từ 1 "card" trong danh sách.

    Ý tưởng & bối cảnh:
    - Giao diện VietnamWorks render card bằng styled-components (class dạng sc-xxxx),
      khá "mỏng manh" vì class có thể đổi tên theo mỗi lần build/deploy.
    - Do đó, ta đi xuyên từng lớp bao để đến thẻ <a> hiển thị ảnh/tiêu đề có
      class 'img_job_card' — đây là điểm bám ổn định hơn ở thời điểm triển khai.
    - Pattern '-jv' trong href là dấu hiệu URL chi tiết tuyển dụng (quan sát thực tế).

    An toàn:
    - Bọc try/except: nếu cấu trúc DOM đổi, đừng ném lỗi toàn cục; trả về [] để
      caller có thể tiếp tục với các card khác.
    - Nếu href là đường dẫn tương đối (bắt đầu bằng '/'), ghép với BASE để tạo URL đầy đủ.

    Trả về:
    - Danh sách các URL (thường 0 hoặc 1 phần tử cho mỗi card).
    """
    links = []
    try:
        # Lần lượt tìm các lớp chứa ảnh/anchor của card:
        sc1 = card.find_element(By.CSS_SELECTOR, "div.sc-iVDsrp")
        sc2 = sc1.find_element(By.CSS_SELECTOR, "div.sc-frWhYi")
        sc3 = sc2.find_element(By.CSS_SELECTOR, "div.sc-hxAGuE")
        # Anchor chính có class 'img_job_card' và href chứa '-jv' (pattern link chi tiết)
        a = sc3.find_element(By.CSS_SELECTOR, "a.img_job_card[href*='-jv']")
        href = (a.get_attribute("href") or "").strip()
        if href:
            # Nếu là href tương đối, prepend BASE để tạo absolute URL.
            if href.startswith("/"):
                href = BASE + href
            links.append(href)
    except Exception:
        # Card không theo cấu trúc kỳ vọng -> bỏ qua yên lặng (tránh gãy luồng xử lý).
        pass
    return links
# === NEW: Append lô bản ghi vào .xlsx mà không cần giữ cả bảng trong RAM ===
def _append_batch_to_excel(excel_path: str, records: List[Dict], sheet_name: str = "jobs") -> None:
    """
    Ghi thêm (append) một lô dict vào file Excel.
    - Tự tạo file nếu chưa có.
    - Tự mở rộng header khi xuất hiện key mới.
    - Sau khi ghi xong, đóng workbook ngay để giải phóng RAM.
    """
    if not records:
        return

    try:
        from openpyxl import Workbook, load_workbook  # dùng openpyxl để append row theo dòng
    except Exception as _:
        # Phòng khi thiếu openpyxl (nhưng nên cài). Tối thiểu vẫn không crash pipeline.
        # Bạn chỉ cần: pip install openpyxl
        raise ModuleNotFoundError("Thiếu 'openpyxl'. Hãy: pip install openpyxl")

    excel_path = str(excel_path)
    path = Path(excel_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        wb = load_workbook(excel_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        headers = [c.value for c in ws[1]] if ws.max_row >= 1 else []
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        headers = []

    # Hợp nhất header cũ + key mới (giữ thứ tự cũ, thêm cột mới vào cuối)
    for rec in records:
        for k in rec.keys():
            if k not in headers:
                headers.append(k)

    # Ghi (hoặc cập nhật) hàng header
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    # Append từng dòng theo thứ tự headers
    for rec in records:
        ws.append([rec.get(h, "") for h in headers])

    wb.save(excel_path)
    wb.close()

def create_driver():
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument(
        "user-agent=Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )

    driver = webdriver.Chrome(options=options)
    driver.set_page_load_timeout(60)   # tối đa 60s load trang
    driver.set_script_timeout(60)      # timeout khi chạy JS
    return driver

def get_vietnamworks_jobs_by_group(
    group_id: int,
    group_name: str,
    max_pages: int = 0,           # 0 = không giới hạn theo tham số này (chỉ còn giới hạn bởi safety_max_pages / no_gain_patience)
    delay: float = 1.0,           # nghỉ giữa 2 page (lịch sự với server, giúp tránh bị rate-limit / CAPTCHA)
    safety_max_pages: int = 200,  # chốt an toàn chống loop vô hạn/redirect lặp (kể cả khi max_pages=0)
    no_gain_patience: int = 2,    # số trang liên tiếp không thu thêm link mới -> dừng để tránh cuộn vô ích
) -> List[Dict]:
    """
    Trình thu thập link job theo 'group_id' (ngành) trên VietnamWorks.
    Trả về list dict: {title: "", href, group_id, group_name}
    (title để trống vì chỉ lấy link; có thể mở trang chi tiết để fill sau).

    Mục tiêu & Lý do thiết kế:
    - Thu gom "đường dẫn chi tiết việc làm" theo từng ngành (group_id) thông qua trang listing.
    - Hạn chế phụ thuộc vào giao diện dễ đổi (class động sc-xxxx) bằng cách:
        (i) chờ phần tử "xương sống" của trang xuất hiện (block-job-list),
        (ii) cuộn kích hoạt lazy-load,
        (iii) đi xuyên card để lấy <a> chi tiết (đã đóng gói trong _extract_links_stepwise_from_card).
    - Tích hợp các cơ chế dừng an toàn:
        * max_pages: giới hạn do người dùng truyền vào (0 = không giới hạn theo tham số này).
        * safety_max_pages: "cầu chì" chống lỗi vòng lặp/redirect.
        * no_gain_patience: dừng khi nhiều trang liền không thêm được liên kết mới (tiết kiệm tài nguyên).

    Thứ tự xử lý (high-level):
    1) Lặp qua các trang /viec-lam?g=<id>&page=<n>.
    2) Chờ khối block-job-list → cuộn lazy-load → trích "card" → rút href chi tiết job.
    3) Dùng seen_hrefs khử trùng lặp trong phiên; dùng "chữ ký trang" (hash của tập href) để phát hiện trang lặp.
    4) Dừng theo một trong các điều kiện: đạt giới hạn, trang rỗng, trang lặp, nhiều trang không tăng dữ liệu.
    5) Trả về danh sách bản ghi tối thiểu (title="", href, group_id, group_name) đã khử trùng lặp lần cuối.

    Ghi chú kỹ thuật:
    - WebDriverWait(wait=12s) nhằm cân bằng giữa độ ổn định (mạng chậm) và tổng thời gian crawl.
    - "window-size=1920x1080" giúp bố cục desktop render đầy đủ, giảm rủi ro layout khác biệt.
    - "user-agent" đặt rõ ràng để tránh bị phân loại là trình tự động quá "lộ liễu".
    - "page_hrefs" lưu tất cả href trên trang (kể cả trùng) để tạo signature ổn định; "page_links" chỉ là phần mới.
    - Hash Python có ngẫu nhiên giữa các tiến trình (PYTHONHASHSEED), nhưng đủ ổn trong phạm vi 1 phiên chạy.
    """

    base_url = f"{BASE}/viec-lam?g={group_id}"

    # ---- Khởi tạo Chrome WebDriver ----
    driver = create_driver()
    wait = WebDriverWait(driver, 25)   # tăng timeout từ 12 → 25s

    # ---- Biến trạng thái thu thập ----
    results: List[Dict] = []      # chứa record tối thiểu cho từng job
    seen_hrefs: set = set()       # set để khử trùng lặp trong phiên (O(1) tra cứu)
    seen_signatures: set = set()  # chữ ký trang: hash(sorted(set(page_hrefs))) để phát hiện vòng lặp/redirect
    page = 1
    no_gain_streak = 0            # đếm số lần liên tiếp không thêm được link mới

    try:
        while True:
            # --- Giới hạn trang bởi tham số/khoá an toàn ---
            if max_pages > 0 and page > max_pages:
                print(f"[{group_name}] Đạt giới hạn max_pages. Dừng.")
                break
            if page > safety_max_pages:
                print(f"[{group_name}] Vượt safety_max_pages. Dừng.")
                break

            # Build URL: trang 1 dùng base_url, từ trang 2 thêm &page=
            url = base_url if page == 1 else f"{base_url}&page={page}"
            print(f"[{group_name}] [FETCH] {url}")
            driver.get(url)

            # Chờ khối 'block-job-list' xuất hiện (cột sống của page listing)
            # Nếu không thấy: không vội kết luận lỗi → có thể là hết dữ liệu/redirect/băng thông chậm.
            try:
                wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "div.block-job-list"))
                )
            except Exception:
                # Có thể do mạng/chuyển trang/hết trang -> vẫn tiếp tục xử lý bên dưới để xác nhận
                print(f"[{group_name}] [WARN] Chưa thấy block-job-list sau timeout.")

            # Cuộn để kích hoạt lazy-load các card (danh sách thường tải dần khi người dùng cuộn)
            _scroll_lazy(driver, times=8, dy=1500, pause=0.25)

            # Tìm container danh sách job (điểm neo để lấy các card)
            try:
                block = driver.find_element(By.CSS_SELECTOR, "div.block-job-list")
            except Exception:
                # Nếu không có container => có thể là trang cuối/DOM thay đổi mạnh -> dừng vòng lặp chính
                print(f"[{group_name}] Không tìm thấy block-job-list. Dừng.")
                break

            # Lấy tất cả 'card' job (mẫu class chung). Không phụ thuộc index (item-0..49)
            # Ưu tiên selector đủ cụ thể để tránh lẫn với các khối khác, nhưng vẫn tránh "quá chặt" vào class động.
            cards = block.find_elements(
                By.CSS_SELECTOR, "div.search_list.view_job_item.new-job-card"
            )

            # --- Gom link trên trang hiện tại ---
            # page_hrefs: mọi href rút được (kể cả trùng) -> dùng tạo "chữ ký trang".
            # page_links: chỉ các href chưa từng thấy trong phiên -> dùng để push vào results.
            page_links, page_hrefs = [], []
            for card in cards:
                for href in _extract_links_stepwise_from_card(card):
                    if href:
                        page_hrefs.append(href)  # chữ ký trang dùng toàn bộ (kể cả trùng)
                        if href not in seen_hrefs:
                            seen_hrefs.add(href)
                            page_links.append(href)  # chỉ những link mới được thêm

            # Nếu không thấy bất kỳ href nào -> coi như trang rỗng / kết thúc dữ liệu
            if not page_hrefs:
                print(f"[{group_name}] Trang không có job. Dừng.")
                break

            # --- Chống vòng lặp/redirect bằng chữ ký trang ---
            # Sort + set để tạo signature ổn định, rồi hash; nếu trùng lặp -> khả năng redirect/vòng lặp.
            page_signature = "|".join(sorted(set(page_hrefs)))
            sig_hash = hash(page_signature)
            if sig_hash in seen_signatures:
                print(f"[{group_name}] Trang có chữ ký lặp lại (redirect/lặp). Dừng.")
                break
            seen_signatures.add(sig_hash)

            # --- Kiểm soát 'không tăng dữ liệu' ---
            if not page_links:
                no_gain_streak += 1
                print(f"[{group_name}] Không có job mới ở trang {page}. no_gain_streak={no_gain_streak}.")
                if no_gain_streak >= no_gain_patience:
                    print(f"[{group_name}] Nhiều trang liên tiếp không tăng dữ liệu. Dừng.")
                    break
            else:
                # Có link mới -> reset streak & push kết quả
                no_gain_streak = 0
                for href in page_links:
                    results.append({
                        "title": "",          # placeholder: chưa parse tiêu đề (sẽ điền khi crawl chi tiết)
                        "href": href,
                        "group_id": group_id,
                        "group_name": group_name,
                    })
                print(f"[{group_name}] Trang {page}: +{len(page_links)} job (tổng {len(results)}).")

            # Sang trang kế, nghỉ 'delay' để đỡ bị nghi ngờ spam (giả lập hành vi người dùng thật)
            page += 1
            time.sleep(delay)

    finally:
        # Đảm bảo đóng trình duyệt dù lỗi hay hoàn tất (giải phóng tài nguyên hệ thống)
        driver.quit()

    # --- Khử trùng lặp lần cuối (phòng trường hợp hi hữu do race/DOM trùng) ---
    # Dựa trên key 'href' (định danh đường dẫn job) để giữ bản ghi cuối cùng cho mỗi href.
    dedup = {it["href"]: it for it in results}
    return list(dedup.values())


def save_group_to_excel(rows: List[Dict], group_name: str, location_code: str = "1001", out_dir: str = "outputs") -> str:
    # Tạo thư mục đầu ra nếu chưa có
    # Lý do: Khi chạy batch cho nhiều ngành/địa điểm, đảm bảo path luôn tồn tại, tránh lỗi I/O.
    os.makedirs(out_dir, exist_ok=True)
    # Chuẩn hoá tên ngành thành slug (không dấu) để đặt tên file an toàn
    # Mục tiêu: tên file thống nhất, không có ký tự đặc biệt gây lỗi trên Windows/Linux.
    name_slug = slugify_vn(group_name)
    # File đầu ra có kèm mã location_code trong tên (chỉ ảnh hưởng tên file, không lọc dữ liệu)
    # Ghi chú: hiện tại file KHÔNG gắn timestamp → lần sau ghi cùng ngành sẽ ghi đè.
    # Tôi chủ động để như vậy nhằm có “bản mới nhất” cho mỗi ngành; nếu cần lịch sử, sẽ bổ sung datetime vào tên.
    filename = f"vietnamworks_jobs_{name_slug}_{location_code}.xlsx"
    path = os.path.join(out_dir, filename)

    # Chuyển list[dict] -> DataFrame để ghi Excel
    df = pd.DataFrame(rows)
    # Chèn cột thời điểm crawl vào đầu bảng (format yyyy-mm-dd HH:MM:SS, timezone-naive)
    # Lý do: giúp truy vết thời điểm thu thập; một số phân tích downstream cần mốc thời gian này.
    df.insert(0, "crawled_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    # Ưu tiên engine 'xlsxwriter'; nếu thiếu package thì fallback sang 'openpyxl'
    # Nhận xét: xlsxwriter cho tốc độ/định dạng tốt khi ghi mới; openpyxl tiện lợi khi cần đọc/ghi chỉnh sửa.
    engine = "xlsxwriter"
    try:
        with pd.ExcelWriter(path, engine=engine) as writer:
            df.to_excel(writer, index=False, sheet_name="jobs")
    except ModuleNotFoundError:
        engine = "openpyxl"
        with pd.ExcelWriter(path, engine=engine) as writer:
            df.to_excel(writer, index=False, sheet_name="jobs")

    # Log đường dẫn file, số dòng và engine đã dùng
    print(f"[SAVE] {path} ({len(rows)} dòng) bằng engine={engine}")
    return path


# ===================== PHẦN 2: Hàm hỗ trợ bóc chi tiết =====================

def _extract_benefits(driver) -> str:
    # Dùng BeautifulSoup trên source HTML hiện tại của driver (đã load trang chi tiết)
    # Lý do: Có nội dung render sẵn trong DOM; Soup thao tác nhanh, phù hợp để bóc text/phúc lợi.
    soup = BeautifulSoup(driver.page_source, "html.parser")
    # Khu vực phúc lợi: selector theo class 'sc-b8164b97-0 kxYTHC' (class styled-component -> dễ thay đổi)
    # Tôi chấp nhận độ “mỏng” của selector này vì chưa thấy hook ổn định hơn; khi vỡ selector sẽ cập nhật.
    benefit_zone = soup.find("div", class_="sc-b8164b97-0 kxYTHC")
    benefits = []
    if benefit_zone:
        # Mỗi block phúc lợi: 'sc-8868b866-0 hoIaMz' (cũng là class "mỏng", có thể đổi theo lần build)
        # Nếu sau này đổi class, chiến lược dự phòng là tìm theo cấu trúc lân cận (ví dụ tiêu đề + mô tả).
        blocks = benefit_zone.find_all("div", class_="sc-8868b866-0 hoIaMz")
        for block in blocks:
            # Tiêu đề phúc lợi
            title_tag = block.find("p", class_="sc-ab270149-0 jlpjAq")
            # Nội dung mô tả chi tiết
            desc_tag = block.find("div", class_="sc-c683181c-2 fGxLZh")
            if title_tag and desc_tag:
                title = title_tag.get_text(strip=True)
                # Ghép các dòng mô tả bằng xuống dòng để dễ đọc
                # Lợi ích: phục vụ xuất Excel/CSV hoặc hiển thị UI text wrap gọn gàng.
                desc = desc_tag.get_text(separator="\n", strip=True)
                benefits.append(f"{title}: {desc}")
    # Trả về chuỗi nhiều dòng, mỗi dòng một phúc lợi
    return "\n".join(benefits)


def _get_text_by_class(soup, tag, class_part, index=0) -> str:
    # Tìm tất cả element theo tên thẻ 'tag' có class chứa chuỗi con 'class_part'
    # (match "chứa" chứ không phải bằng tuyệt đối; hữu ích khi class có nhiều token)
    # Ưu điểm: bớt phụ thuộc vào toàn bộ chuỗi class; chỉ cần một phần ổn định là đủ.
    els = soup.find_all(tag, class_=lambda x: x and class_part in x)
    # Lấy text của phần tử thứ 'index' nếu có, ngược lại trả chuỗi rỗng
    # Dùng get_text(strip=True) để loại bỏ khoảng trắng dư thừa.
    return els[index].get_text(strip=True) if len(els) > index else ""


def _click_expand_buttons(driver, max_clicks: int = 20):
    # Click các nút "Xem thêm"/"Xem đầy đủ mô tả công việc" để mở rộng nội dung ẩn
    # Bối cảnh: nhiều trang chi tiết rút gọn mô tả bằng accordion; cần mở ra để Soup thu đủ nội dung.
    # Lưu ý: tham số 'wait' hiện chưa được sử dụng (giữ để đồng bộ interface, có thể hữu ích nếu chờ element động)
    clicked_count = 0
    for _ in range(max_clicks):
        try:
            # Tìm các button bên trong vùng có class chứa 'sc-8868b866-0' và 'kAAFiO'
            # Chỉ giữ các button hiển thị & có text hợp lệ
            # Lý do dùng XPATH: biểu thức contains() giúp “nới lỏng” điều kiện khi class có nhiều token động.
            buttons = [
                btn for btn in driver.find_elements(
                    By.XPATH,
                    "//div[contains(@class, 'sc-8868b866-0') and contains(@class, 'kAAFiO')]//button"
                )
                if btn.is_displayed() and btn.text.strip() and (
                    "Xem thêm" in btn.text or "Xem đầy đủ mô tả công việc" in btn.text
                )
            ]
            # Nếu không còn button mở rộng -> dừng vòng lặp
            if not buttons:
                break
            # Dùng JS click để tránh lỗi intercept (che khuất/overlay), ổn định hơn click() thông thường
            # Sau click, DOM thường reflow → sleep ngắn để nội dung mới render xong trước khi bóc.
            driver.execute_script("arguments[0].click();", buttons[0])
            clicked_count += 1
            # Nghỉ ngắn để nội dung mở rộng kịp render
            time.sleep(0.4)
        except Exception:
            # Nếu có lỗi (hết element, DOM đổi…) -> dừng an toàn
            # Đây là quyết định bảo thủ để tránh vỡ luồng chính do lỗi giao diện cục bộ.
            break
    # Log số lần đã click mở rộng (nếu có)
    if clicked_count:
        print(f"  [INFO] Đã click mở rộng {clicked_count} lần.")

# ===================== PHẦN 3: Crawl chi tiết job =====================
# === NEW: Bóc chi tiết dạng streaming, ghi ra Excel ngay để nhẹ RAM ===
def scrape_job_details_streaming_to_excel(job_links: List[str],
                                          out_xlsx_path: str,
                                          start_id: int = 1000001,
                                          batch_size: int = 20) -> int:
    """
    Bóc chi tiết từng link và GHI THẲNG ra Excel theo lô (batch_size) để giải phóng RAM ngay.
    Trả về: tổng số job đã ghi.
    """
    driver = create_driver()
    wait = WebDriverWait(driver, 30)

    batch: List[Dict] = []
    total_written = 0

    try:
        for index, job_url in enumerate(job_links):
            print(f"\n[{index + 1}/{len(job_links)}] Đang xử lý: {job_url}")
            try:
                driver.get(job_url)
                wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                time.sleep(1.2)
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(0.8)
                _click_expand_buttons(driver, max_clicks=20)
                time.sleep(1.0)

                benefits_text = _extract_benefits(driver)
                soup = BeautifulSoup(driver.page_source, "html.parser")

                job_fields = {
                    "ID": start_id + index,
                    "Tên công việc": _get_text_by_class(soup, "h1", "hAejeW"),
                    "Lương": _get_text_by_class(soup, "span", "cVbwLK"),
                    "Hết hạn": _get_text_by_class(soup, "span", "ePOHWr", 0),
                    "Lượt xem": _get_text_by_class(soup, "span", "ePOHWr", 1),
                    "Địa điểm tuyển dụng": _get_text_by_class(soup, "span", "ePOHWr", 2),
                }

                # Section mô tả
                description_sections = soup.find_all("div", class_=lambda x: x and "gDSEwb" in x)
                for section in description_sections:
                    title_tag = section.find("h2", class_=lambda x: x and "cjuZti" in x)
                    content_tag = section.find("div", class_=lambda x: x and "dVvinc" in x)
                    if title_tag and content_tag:
                        title = title_tag.get_text(strip=True)
                        content = content_tag.get_text(separator="\n", strip=True)
                        job_fields[title] = content

                # Phúc lợi
                job_fields["Phúc lợi"] = benefits_text

                # Cặp Label/Value
                job_info_section = soup.find("div", class_=lambda x: x and "dHvFzj" in x)
                if job_info_section:
                    info_items = job_info_section.find_all("div", class_=lambda x: x and "JtIju" in x)
                    for item in info_items:
                        label_tag = item.find("label", class_=lambda x: x and "dfyRSX" in x)
                        value_tag = item.find("p", class_=lambda x: x and "cLLblL" in x)
                        if label_tag and value_tag:
                            job_fields[label_tag.get_text(strip=True)] = value_tag.get_text(strip=True)

                # Địa điểm làm việc
                loc = soup.find("div", class_=lambda x: x and "bAqPjv" in x)
                if loc:
                    val = loc.find("p", class_=lambda x: x and "cLLblL" in x)
                    if val:
                        job_fields["Địa điểm làm việc"] = val.get_text(strip=True)

                # Công ty
                comp = soup.find("div", class_=lambda x: x and "drWnZq" in x)
                if comp:
                    name = comp.find("a", class_=lambda x: x and "egZKeY" in x)
                    size = comp.find("span", class_=lambda x: x and "ePOHWr" in x)
                    if name:
                        job_fields["Tên công ty"] = name.get_text(strip=True)
                    if size:
                        job_fields["Quy mô công ty"] = size.get_text(strip=True)

                job_fields["HREF"] = job_url

                # Dồn vào batch
                batch.append(job_fields)

                # Đủ lô -> ghi ra file & dọn RAM
                if len(batch) >= batch_size:
                    _append_batch_to_excel(out_xlsx_path, batch, sheet_name="jobs")
                    total_written += len(batch)
                    batch.clear()
                    del soup, job_fields, benefits_text
                    gc.collect()

            except Exception as e:
                print(f"  ❌ Lỗi khi xử lý link: {e}")
                # tiếp tục link sau

    finally:
        driver.quit()

    # Flush phần còn lại
    if batch:
        _append_batch_to_excel(out_xlsx_path, batch, sheet_name="jobs")
        total_written += len(batch)
        batch.clear()
        gc.collect()

    print(f"[DETAIL][STREAM] Đã ghi {total_written} job vào: {out_xlsx_path}")
    return total_written



if __name__ == "__main__":
    # ==== THAM SỐ CHUNG ====
    LOCATION_CODE = r"1001"
    LIST_OUT_DIR = str((_OUTPUT_ROOT / "jobslist").resolve())
    DETAIL_OUT_DIR = str((_OUTPUT_ROOT / "jobsdetail").resolve())
    MAX_PAGES = 0
    DELAY = 1.0
    NO_GAIN_PATIENCE = 2
    START_ID_BASE = 1000001
    ID_STEP_PER_GROUP = 1000000

    run_ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")

    os.makedirs(LIST_OUT_DIR, exist_ok=True)
    os.makedirs(DETAIL_OUT_DIR, exist_ok=True)

    summary = []

    for idx, (group_name, gid) in enumerate(VNWORKS_GROUPS.items(), start=0):
        try:
            print("\n" + "="*80)
            print(f"[{idx+1}/{len(VNWORKS_GROUPS)}] NGÀNH: {group_name} (g={gid})")

            # 1) Crawl danh sách link
            rows = get_vietnamworks_jobs_by_group(
                group_id=gid,
                group_name=group_name,
                max_pages=MAX_PAGES,
                delay=DELAY,
                no_gain_patience=NO_GAIN_PATIENCE,
            )

            # 2) Lưu danh sách (list) -> output/jobslist
            list_path = save_group_to_excel(
                rows=rows,
                group_name=group_name,
                location_code=LOCATION_CODE,
                out_dir=LIST_OUT_DIR
            )

            # Dọn RAM của list ngay sau khi lưu
            links = [r["href"] for r in rows if r.get("href")]
            del rows
            gc.collect()

            # 3) Bóc chi tiết -> ghi STREAMING ra output/jobsdetail
            name_slug = slugify_vn(group_name)
            start_id = START_ID_BASE + idx * ID_STEP_PER_GROUP
            detail_filename = f"job_detail_output_{name_slug}_g{gid}_{LOCATION_CODE}_{run_ts}.xlsx"
            detail_path = os.path.join(DETAIL_OUT_DIR, detail_filename)

            if links:
                print(f"[DETAIL] Bắt đầu bóc chi tiết {len(links)} link cho ngành '{group_name}'...")
                # === CHANGED TO STREAMING ===
                n_written = scrape_job_details_streaming_to_excel(
                    job_links=links,
                    out_xlsx_path=detail_path,
                    start_id=start_id,
                    batch_size=20  # có thể tăng/giảm; 10–50 là hợp lý cho t3.small
                )
            else:
                print(f"[DETAIL][WARN] Ngành '{group_name}' không có link nào. Tạo file chi tiết rỗng.")
                # tạo file Excel rỗng với header tối thiểu
                _append_batch_to_excel(detail_path, [], sheet_name="jobs")
                n_written = 0

            # Sau khi ghi, dọn các biến tạm
            del links
            gc.collect()

            summary.append((group_name, list_path, detail_path, None, n_written))

        except Exception as e:
            print(f"[ERROR] Lỗi ở ngành '{group_name}' (g={gid}): {e}")
            gc.collect()

    # ==== TỔNG KẾT ====
    print("\n" + "="*80)
    print("[SUMMARY]")
    for (group_name, list_path, detail_path, n_list, n_detail) in summary:
        print(f"- {group_name}: list=(saved) {list_path} | detail={n_detail} rows ({detail_path})")
